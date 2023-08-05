import PySimpleGUI as sg
import openpyxl
import os

def executa_app():
    try:
        caminho_app = os.path.abspath(__file__).replace('\\','/').replace('.py','.exe')
        os.startfile(caminho_app)
    except FileNotFoundError:
        os.system('python create_table.py')
    except FileNotFoundError:
        print('Nenhum arquivo ou diretorio encontrado')

sg.theme('Topanga')

def window_tabela():
    layout = [
        [sg.Text('Preencha os campos abaixo para criar uma nova Tabela')],
        [sg.Text(f'Digite o nome da planilha que será criada{os.linesep}Crie quantas quiser')],
        [sg.Input(key='sheet'), sg.Button('Criar Planilha')],
        [sg.Text('Digite o nome da planilha que irá receber os dados')],
        [sg.Text(key='visualizar planilhas')],
        [sg.Input(key='sheet escolha'), sg.Button('selecionar', disabled=True, size=(10,1))],
        [sg.Text(key='resposta selecao')],
        [sg.Text('Digite o nome das colunas que serão criadas separados por \",\"')],
        [sg.Input(key='colunas'), sg.Button('Criar Colunas', disabled=True, size=(10,1))],
        [sg.Text(f'Digite os valores que serão inseridos nas colunas separados por \",\"{os.linesep}Adicione quantas linhas quiser')],
        [sg.Text(key='vizualizar coluna')],
        [sg.Input(key='valores'), sg.Button('adicionar linha', disabled=True, size=(10,1))],
        [sg.Text(key='vizualizar valor')],
        [sg.Text('Digite o nome da Tabela')],
        [sg.Input(key='nome tabela'), sg.Button('Criar Tabela', disabled=True, size=(10,1))],
    ]
    workbook = openpyxl.Workbook()
    del workbook['Sheet']

    return sg.Window('Criando Tabela', layout, finalize=True), workbook

window, workbook = window_tabela()

def main():
    while True:
        event, values = window.read()
        if event == sg.WIN_CLOSED:
            break
        elif event == 'Criar Planilha':
            window['selecionar'].update(disabled=False)
            window['Criar Tabela'].update(disabled=False)
            workbook.create_sheet(values['sheet'])
            window['visualizar planilhas'].update(f'Planilhas: {workbook.sheetnames}', text_color='white')
            window['sheet'].update('')
        
        elif event == 'selecionar':
            try:
                window['Criar Colunas'].update(disabled=False)
                sheet = workbook[values['sheet escolha']]
                window['resposta selecao'].update(f'Planinha Selecionada: {values["sheet escolha"]}', text_color='white')
                window['sheet escolha'].update('')
            except KeyError:
                sg.popup_error('Digite o nome de planinha uma que você tenha criado', title='Planilha Errada')

        elif event == 'Criar Colunas':
            if ',' and not '.' in values['colunas']:
                window['adicionar linha'].update(disabled=False)
                sheet.append(values['colunas'].split(','))
                window['Criar Colunas'].update(disabled=True)
                window['vizualizar coluna'].update(f"{values['colunas'].split(',')}", text_color='white')
                window['colunas'].update('')
            else:
                sg.popup_error('Digite apenas valores separados por \",\"', title='Erro')

        elif event == 'adicionar linha':
            if ',' and not '.' in values['valores']:
                sheet.append(values['valores'].split(','))
                window['valores'].update('')
                window['vizualizar valor'].update(f'Ultimo valor adicionado: {values["valores"]}', text_color='white')
            else:
                sg.popup_error('Digite apenas valores separados por \",\"', title='Erro')
                
        elif event == 'Criar Tabela':
            if not os.path.exists(f'{values["nome tabela"]}.xlsx'):
                workbook.save(f'{values["nome tabela"]}.xlsx')
                resposta = sg.popup(f'Tabela {values["nome tabela"]}.xlsx criada com sucesso!!{os.linesep}Criar outra Tabela',custom_text=('Sim','Não'), title='Sucesso')
                if resposta == 'Sim':
                    window.close()
                    executa_app()
                else:
                    break
            else:
                sg.popup_error('Essa Tabela já existe, Digite outro nome!', title='Tabela existente')

if __name__ == '__main__':
    main()